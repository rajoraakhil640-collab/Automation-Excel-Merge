import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import argparse
import os

def merge_files(emp_file, sal_file, output_base):
    try:
      #Check file existence
        if not os.path.exists(emp_file):
            raise FileNotFoundError(f"{emp_file} not found.")
        if not os.path.exists(sal_file):
            raise FileNotFoundError(f"{sal_file} not found.")

       #Read Excel file
        emp_df = pd.read_excel(emp_file, engine="openpyxl")
        sal_df = pd.read_excel(sal_file, engine="openpyxl")
       #Merge both file using Employee_Id
        merged_df = pd.merge(emp_df, sal_df, on="Employee_ID", how="left")

        # Add 'Status' Column
        merged_df["Status"] = merged_df["Salary"].notna().map({True: "Matched", False: "Not Found"})

        # Summary
        total = len(merged_df)
        matched = (merged_df["Status"] == "Matched").sum()
        not_found = (merged_df["Status"] == "Not Found").sum()

       #Merge Summary
        print("\n=== MERGE SUMMARY ===")
        print(f"Total Employees : {total}")
        print(f"Matched Records : {matched}")
        print(f"Unmatched Records : {not_found}")

      # Add timestamp to outfile filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_output = f"{output_base}_{timestamp}.xlsx"
        csv_output = f"{output_base}_{timestamp}.csv"

        #Save as Excel and CSV
        with pd.ExcelWriter(excel_output, engine="openpyxl") as writer:
         merged_df.to_excel(writer, index=False)
        print(f"\nOutput saved as:")
        print(f"   Excel → {excel_output}")
        print(f"   CSV   → {csv_output}")

       # Highlight unmatched rows in Excel
        wb = load_workbook(excel_output)
        ws = wb.active
        red_fill = PatternFill(fill_type="solid", fgColor="FF9999")
        status_col = merged_df.columns.get_loc("Status") + 1

        for row in ws.iter_rows(min_row=2):
            if row[status_col - 1].value == "Not Found":
                for cell in row:
                    cell.fill = red_fill

        wb.save(excel_output)
        print(f"Unmatched rows highlighted in red inside {excel_output}")

    except FileNotFoundError as e:
        print(f"Error: {e}")
    except KeyError as e:
        print(f"Missing column: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")

# Command-Line Interface
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Merge employee and salary Excel files (VLOOKUP style).")

    parser.add_argument(
        "employees_file",
        nargs="?",
        default="employees.xlsx",
        help="Path to employees.xlsx file (default: employees.xlsx)"
    )
    parser.add_argument(
        "salaries_file",
        nargs="?",
        default="salaries.xlsx",
        help="Path to salaries.xlsx file (default: salaries.xlsx)"
    )
    parser.add_argument(
        "-o", "--output",
        help="Base name for output files",
        default="Employee_salary_report.xlsx"
    )

    args = parser.parse_args()
    merge_files(args.employees_file, args.salaries_file, args.output)

